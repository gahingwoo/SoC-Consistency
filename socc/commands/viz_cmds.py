"""socc viz <subcommand> — visualization commands."""

from __future__ import annotations

import click

from socc.commands._shared import (
    ALL_SOC_CHOICES, FuzzySoCType, build_registry, auto_detect_soc,
    build_sample_model, parse_dts_file, parse_dts_cached, Checker, Path, Optional,
)


@click.group("viz")
def viz_group():
    """Visualization: topology graph, BGA pinmap, power-rail sequence."""


# ── viz topology ──────────────────────────────────────────────────────────────

@viz_group.command("topology")
@click.argument("dts_file", type=click.Path(exists=True), required=False)
@click.option("--soc", type=FuzzySoCType(ALL_SOC_CHOICES), metavar="SOC", default=None)
@click.option("-o", "--output", default=None, metavar="FILE")
@click.option("--demo", is_flag=True, help="Use built-in sample model.")
@click.option("--no-open", is_flag=True,
              help="Do not open the HTML in the browser after generation.")
@click.option("--with-violations", is_flag=True,
              help="Run checker and annotate violations onto the graph.")
def viz_topology(dts_file: Optional[str], soc: Optional[str], output: Optional[str],
                 demo: bool, no_open: bool, with_violations: bool):
    """Generate an interactive HTML hardware topology graph.

    \b
    Example:
        socc viz topology board.dts --with-violations
    """
    import subprocess, sys
    from socc.visualize.topology import render_topology_html

    if demo:
        model = build_sample_model("rk3588")
        soc_name = "rk3588"
    else:
        if not dts_file:
            click.echo("Error: specify a DTS file or use --demo", err=True)
            raise SystemExit(1)
        soc_name = soc if (soc and soc != "auto") else auto_detect_soc(Path(dts_file).name)
        try:
            model = parse_dts_file(dts_file, soc_name)
        except Exception as e:
            click.echo(f"Error: {e}", err=True)
            raise SystemExit(1)

    violations = []
    if with_violations:
        checker = Checker(build_registry())
        violations = checker.check(model, soc_name)

    out_path = output or f"{soc_name}-topology.html"
    title = f"SoC Topology — {soc_name}"
    if dts_file:
        title += f" ({Path(dts_file).name})"

    try:
        result = render_topology_html(model=model, output_path=out_path,
                                      violations=violations, title=title)
    except ImportError as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)

    viol_note = f" ({len(violations)} violation(s) annotated)" if violations else ""
    click.echo(f"Topology graph written to: {result}{viol_note}")

    if not no_open:
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", result])
            elif sys.platform.startswith("linux"):
                subprocess.Popen(["xdg-open", result])
            elif sys.platform == "win32":
                subprocess.Popen(["start", result], shell=True)
        except Exception:
            pass


# ── viz pinmap ────────────────────────────────────────────────────────────────

@viz_group.command("pinmap")
@click.argument("dts_file", type=click.Path(exists=True), required=False)
@click.option("--soc", "soc_name", type=FuzzySoCType(ALL_SOC_CHOICES), metavar="SOC",
              default=None, required=True,
              help="SoC identifier (must have a 'bga:' section in its YAML).")
@click.option("-o", "--output", default=None, metavar="FILE")
@click.option("--format", "fmt", type=click.Choice(["html", "xlsx", "csv"]),
              default="html", show_default=True,
              help="Output format: html (BGA heatmap), xlsx (Excel matrix), csv.")
@click.option("--no-dts", is_flag=True,
              help="Generate from YAML data only, without DTS cross-reference.")
def viz_pinmap(dts_file: Optional[str], soc_name: str, output: Optional[str],
               fmt: str, no_dts: bool):
    """Generate a pin-assignment matrix for a SoC.

    \b
    Examples:
        socc viz pinmap board.dts --soc rk3588
        socc viz pinmap board.dts --soc rk3588 --format xlsx -o rk3588_pins.xlsx
        socc viz pinmap --soc rk3588 --format csv -o pins.csv
    """
    dts_pinmux = None
    if dts_file and not no_dts:
        try:
            model = parse_dts_cached(dts_file, soc_name)
            dts_pinmux = model.pinmux_config or None
        except Exception as e:
            click.echo(f"Warning: could not parse DTS for cross-reference: {e}", err=True)

    if fmt == "html":
        from socc.visualize.pinmap import render_bga_heatmap
        out_path = output or f"{soc_name}-pinmap.html"
        try:
            result = render_bga_heatmap(soc_name=soc_name, output_path=out_path,
                                        dts_pinmux=dts_pinmux)
            click.echo(f"Pinout heatmap written to: {result}")
        except ValueError as e:
            click.echo(f"Error: {e}", err=True)
            raise SystemExit(1)
    else:
        # xlsx / csv — build pin table from pinmux_config
        _export_pin_table(soc_name, dts_pinmux, output, fmt)


# ── viz power-seq ─────────────────────────────────────────────────────────────

@viz_group.command("power-seq")
@click.argument("dts_file", type=click.Path(exists=True), required=False)
@click.option("--soc", type=FuzzySoCType(ALL_SOC_CHOICES), metavar="SOC", default=None)
@click.option("--color/--no-color", default=None)
@click.option("--demo", is_flag=True, help="Use built-in sample model.")
def viz_power_seq(dts_file: Optional[str], soc: Optional[str],
                  color: Optional[bool], demo: bool):
    """Visualise the power-rail startup sequence as an ASCII timing waveform.

    \b
    Example:
        socc viz power-seq board.dts
    """
    from socc.visualize.power_seq import render_power_sequence
    use_color = color
    if demo:
        model = build_sample_model("rk3588")
        title = "Power Rail Startup Sequence — demo (rk3588)"
    else:
        if not dts_file:
            click.echo("Error: specify a DTS file or use --demo", err=True)
            raise SystemExit(1)
        soc_name = soc if (soc and soc != "auto") else auto_detect_soc(Path(dts_file).name)
        try:
            model = parse_dts_file(dts_file, soc_name)
        except Exception as e:
            click.echo(f"Error: {e}", err=True)
            raise SystemExit(1)
        title = f"Power Rail Startup Sequence — {Path(dts_file).name}"
    click.echo(render_power_sequence(model, use_color=use_color, title=title))


# ── pin-table export helper ───────────────────────────────────────────────────

_PIN_HEADERS = [
    "Pin Name", "GPIO Bank", "Pin Index", "Alt Function",
    "Direction", "Pull", "Drive Strength mA", "Net / Peripheral",
]


def _build_pin_rows(soc_name: str, dts_pinmux: Optional[dict]) -> list:
    """Build a list of row dicts for the pin export table.

    Falls back to an illustrative placeholder row when no pinmux data is
    available so the generated file is never empty.
    """
    rows = []
    if dts_pinmux:
        for pin_name, attrs in sorted(dts_pinmux.items()):
            if isinstance(attrs, dict):
                rows.append({
                    "Pin Name":         pin_name,
                    "GPIO Bank":        attrs.get("bank", ""),
                    "Pin Index":        attrs.get("pin", ""),
                    "Alt Function":     attrs.get("function", ""),
                    "Direction":        attrs.get("direction", ""),
                    "Pull":             attrs.get("bias", ""),
                    "Drive Strength mA": attrs.get("drive_strength", ""),
                    "Net / Peripheral": attrs.get("peripheral", ""),
                })
            else:
                rows.append({
                    "Pin Name": pin_name, "GPIO Bank": "", "Pin Index": "",
                    "Alt Function": str(attrs), "Direction": "", "Pull": "",
                    "Drive Strength mA": "", "Net / Peripheral": "",
                })
    if not rows:
        rows.append({h: f"(no DTS pinmux data for {soc_name})" if h == "Pin Name" else ""
                     for h in _PIN_HEADERS})
    return rows


def _export_pin_table(
    soc_name: str,
    dts_pinmux: Optional[dict],
    output: Optional[str],
    fmt: str,
) -> None:
    rows = _build_pin_rows(soc_name, dts_pinmux)

    if fmt == "csv":
        import csv, io
        out_path = output or f"{soc_name}-pinmap.csv"
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=_PIN_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        Path(out_path).write_text(buf.getvalue())
        click.echo(f"Pin assignment CSV written to: {out_path}")
        return

    # xlsx
    out_path = output or f"{soc_name}-pinmap.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        click.echo(
            click.style(
                "openpyxl is required for XLSX export.  Install it with:\n"
                "  pip install openpyxl",
                fg="yellow"),
            err=True,
        )
        raise SystemExit(1)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"{soc_name} Pin Assignment"

    # ── styles ────────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    even_fill   = PatternFill("solid", fgColor="D6E4F0")   # light blue
    odd_fill    = PatternFill("solid", fgColor="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── header row ────────────────────────────────────────────────────────────
    ws.append(_PIN_HEADERS)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = thin_border
        cell.alignment = center_align
    ws.row_dimensions[1].height = 22

    # ── data rows ─────────────────────────────────────────────────────────────
    for idx, row in enumerate(rows, start=2):
        ws.append([row.get(h, "") for h in _PIN_HEADERS])
        fill = even_fill if idx % 2 == 0 else odd_fill
        for cell in ws[idx]:
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = left_align

    # ── column widths ─────────────────────────────────────────────────────────
    col_widths = [20, 12, 12, 18, 12, 12, 20, 28]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ── freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(out_path)
    click.echo(click.style(f"Pin assignment Excel written to: {out_path}", fg="green"))
